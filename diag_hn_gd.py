# -*- coding: utf-8 -*-
import openpyxl, re, json, sys
sys.path.insert(0, 'src')
import railway_data as rd

BASE = 'C:/Users/cjp15/Desktop/全国客运站点/各省市细分站点/湖南省、广东省、香港特别行政区/'
FILES = ['湖南省补充站点总表.xlsx', '广东省补充站点总表.xlsx']

def clean_station(cell):
    """Extract canonical station name from a 车站名称 or 前/后一站 cell."""
    if cell is None:
        return None
    s = str(cell).strip()
    if s in ('—', '-', '', '终点站', '分岔点'):
        return None
    # directional / non-station text -> no 'X站' at start
    m = re.match(r'^([\u4e00-\u9fff]+站)', s)
    if not m:
        # try to find any 'X站' token
        m2 = re.search(r'([\u4e00-\u9fff]+站)', s)
        if not m2:
            return None
        cand = m2.group(1)
        # if it's a terminal-ish word, return None
        if cand.startswith(('终点站', '分岔点')):
            return None
        return cand
    name = m.group(1)
    # parenthetical standard-name hint: e.g. 大石冬站（大石东站） -> 大石东站
    pm = re.search(r'（([\u4e00-\u9fff]+站)）', s)
    if pm:
        return pm.group(1)
    return name

def parse_excel(fn):
    wb = openpyxl.load_workbook(BASE+fn, data_only=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row+1):
        vals = [ws.cell(r,c).value for c in range(1, 8)]
        name = vals[0]
        if name is None:
            continue
        if str(name).startswith('说明') or str(name).startswith('【'):
            continue  # note / batch-summary rows
        if str(name) == '车站名称':
            continue  # header
        st = clean_station(name)
        if not st:
            continue
        line = vals[1]
        if line is None:
            continue
        line = str(line).strip()
        prov = str(vals[3]).strip() if vals[3] else ''
        prev = clean_station(vals[4])
        nxt = clean_station(vals[5])
        rows.append({'st': st, 'line': line, 'prov': prov, 'prev': prev, 'next': nxt,
                     'raw_prev': vals[4], 'raw_next': vals[5]})
    return rows

all_rows = []
for fn in FILES:
    rows = parse_excel(fn)
    print(f'=== {fn}: parsed {len(rows)} station-rows ===')
    all_rows.extend(rows)

# group by line
from collections import defaultdict
by_line = defaultdict(list)
for r in all_rows:
    by_line[r['line']].append(r)

print(f'\nTOTAL supplement stations: {len(all_rows)}')
print(f'DISTINCT lines: {len(by_line)}')

# Province from prov field
print('\n=== province city keys sample ===')
provcities = set()
for r in all_rows:
    m = re.match(r'(湖南省|广东省|香港特别行政区)([\u4e00-\u9fff]+?)(市|区|县|自治州|特别行政区|新区)$', r['prov'])
    if m:
        provcities.add((m.group(1), m.group(2)+m.group(3)))
for pc in sorted(provcities)[:60]:
    print(' ', pc)

# DB cross-reference
print('\n=== DB cross-reference ===')
db_lines = set(rd.LINE_ORDER.keys())
new_lines = [l for l in by_line if l not in db_lines]
print(f'Lines NOT in DB (fully new): {len(new_lines)}')
for l in sorted(new_lines):
    print('   NEW LINE:', l, '->', len(by_line[l]), 'stations')

existing_lines = [l for l in by_line if l in db_lines]
print(f'\nLines ALREADY in DB: {len(existing_lines)}')
for l in sorted(existing_lines):
    sup_stations = set(x['st'] for x in by_line[l])
    db_seq = rd.LINE_ORDER[l]
    already = sup_stations & set(db_seq)
    print(f'   {l}: sup={len(sup_stations)} db_seq={len(db_seq)} already_in_db={len(already)}')

# Station existence
sup_all = set(r['st'] for r in all_rows)
in_db = sup_all & set(rd.STATION_INFO.keys())
print(f'\nSupplement stations already in STATION_INFO: {len(in_db)}')
print('  ', sorted(in_db)[:40])

# Anchor check: for each supplement station, do prev/next resolve to a station that exists
# in DB or in supplement?
known = set(rd.STATION_INFO.keys()) | sup_all
dangling = []
for r in all_rows:
    for nb in (r['prev'], r['next']):
        if nb and nb not in known:
            dangling.append((r['st'], nb))
print(f'\nPrev/next neighbors NOT in DB nor supplement (need anchor/synth): {len(dangling)}')
for s, nb in dangling[:50]:
    print('   ', s, '->', nb)
