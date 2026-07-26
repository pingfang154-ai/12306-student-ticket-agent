# -*- coding: utf-8 -*-
import openpyxl, re, json, sys
from collections import defaultdict, OrderedDict

PY = "C:/Users/cjp15/.workbuddy/binaries/python/envs/default/Scripts/python.exe"
XLSX = r"C:\Users\cjp15\Desktop\全国客运站点\各省市细分站点\江西省、福建省\江西省、福建省补充站点总表.xlsx"

wb = openpyxl.load_workbook(XLSX, data_only=True)
print("SHEETS:", wb.sheetnames)
ws = wb[wb.sheetnames[0]]

# header row: find the row containing 车站名称
header_row = None
for r in range(1, min(ws.max_row, 30)+1):
    vals = [str(ws.cell(r,c).value).strip() for c in range(1, ws.max_column+1)]
    if any("车站名称" in v for v in vals):
        header_row = r; break
print("HEADER ROW:", header_row)

# map header -> col
hdr = {}
for c in range(1, ws.max_column+1):
    v = ws.cell(header_row, c).value
    if v is not None:
        hdr[str(v).strip()] = c
print("HEADER MAP:", hdr)

def get(row, key):
    if key not in hdr: return ""
    v = ws.cell(row, hdr[key]).value
    return "" if v is None else str(v).strip()

def prov_of(city):
    # 所属省市 like "江西省赣州市" / "福建省三明市"
    for p in ("江西省","福建省"):
        if city.startswith(p): return p
    if city.startswith("江西"): return "江西省"
    if city.startswith("福建"): return "福建省"
    return ""

rows = []
for r in range(header_row+1, ws.max_row+1):
    name = get(r, "车站名称")
    if not name: continue
    rows.append({
        "prov": prov_of(get(r, "所属省市")),
        "city": get(r, "所属省市"),
        "line": get(r, "所属线路"),
        "station": name,
        "seq": get(r, "同线顺序"),
        "prev": get(r, "前一站"),
        "next": get(r, "后一站"),
        "note": get(r, "备注（客运状态）"),
    })

print("TOTAL DATA ROWS:", len(rows))

# group by province
byprov = defaultdict(int)
for x in rows: byprov[x["prov"] or "??"] += 1
print("BY PROVINCE:", dict(byprov))

# distinct stations / lines
stations = OrderedDict()
lines = defaultdict(list)
for x in rows:
    lines[x["line"]].append(x)
    stations.setdefault(x["station"], x)
print("DISTINCT STATIONS:", len(stations))
print("DISTINCT LINES:", len(lines))
for ln in lines:
    print(f"  LINE: {ln}  | n={len(lines[ln])} | cities={sorted({x['city'] for x in lines[ln]})}")

# now cross-check against current DB
import importlib.util as u
spec = u.spec_from_file_location('rd', 'src/railway_data.py')
rd = u.module_from_spec(spec); spec.loader.exec_module(rd)
META = rd.META
print("\n=== CURRENT DB META version =", META.get('version'))
print("DB lines:", len(rd.LINE_ORDER), "stations:", len(rd.STATION_INFO))

existing_stations = set(rd.STATION_INFO.keys())
existing_lines = set(rd.LINE_ORDER.keys())

new_stations = [s for s in stations if s not in existing_stations]
print("NEW STATIONS (not in DB):", len(new_stations))
for s in new_stations:
    print("   +", s, "|", stations[s]["city"], "|", stations[s]["line"])

new_lines = [l for l in lines if l not in existing_lines]
print("NEW LINES (not in DB):", len(new_lines))
for l in new_lines: print("   NEWLINE:", l)

# stations already in DB
already = [s for s in stations if s in existing_stations]
print("ALREADY-IN-DB STATIONS:", len(already), already[:20])

# prev/next that are missing (need anchor)
miss_prevnext = set()
for x in rows:
    for k in ("prev","next"):
        v = x[k]
        if v and v not in existing_stations and v not in stations:
            miss_prevnext.add((x["station"], k, v))
print("PREV/NEXT refs NOT in DB and NOT in this excel (need external anchor):", len(miss_prevnext))
for t in sorted(miss_prevnext): print("   anchor?", t)

# check prev/next vs seq ordering consistency
print("\n=== prev/next vs 同线顺序 sanity (first 15) ===")
for x in rows[:15]:
    print(f"  {x['station']:12} line={x['line'][:18]:18} seq={x['seq']:4} prev={x['prev']:10} next={x['next']:10} city={x['city']}")
