# -*- coding: utf-8 -*-
import openpyxl
XLSX = r"C:\Users\cjp15\Desktop\全国客运站点\各省市细分站点\江西省、福建省\江西省、福建省补充站点总表.xlsx"
wb = openpyxl.load_workbook(XLSX, data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    print("=== SHEET", repr(sn), "dims", ws.dimensions, "max_row", ws.max_row, "max_col", ws.max_column)
    for r in range(1, min(ws.max_row, 8) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        print(r, [str(v)[:16] for v in vals])
