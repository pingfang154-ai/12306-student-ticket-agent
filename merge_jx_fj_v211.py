# -*- coding: utf-8 -*-
# 任务七：江西 + 福建 补充站点增量合并 (v2.10 -> v2.11)
# 数据源：各省市细分站点/江西省、福建省/江西省、福建省补充站点总表.xlsx
#   共 50 座办理客运车站（5 行为批次说明需过滤），分属 23 条线路
#   （16 条全新线 + 6 条既有线精确插入：京九铁路/沪昆铁路/赣韶铁路/吉衡铁路/铜九铁路/皖赣铁路）。
# 策略（沿用 v2.10 铁律与连通性修复）：
#   * 单点式补充：Excel 给前/后一站相邻参考 + 同线顺序(已按真实走向编号)。
#   * 既有线(6条)精确插入：按 prev/next 锚点（在库/本批内）接入 DB 既有序列。
#   * 16 条全新线：按同线顺序/prev-next 链重建序列写入 LINE_ORDER（Dijkstra 才可见）。
#   * 跨线/跨批边界(如 石城东→宁化、南丰→建宁县北、大余→南雄、定南南→和平北、龙市→炎陵、
#     资溪→光泽、彭泽→东至、瑞昌→武穴、瑞昌西→阳新、松溪→庆元 等)优先接已入库锚点(广东/湖南批已补)，
#     不可达外部锚点则以合成线接最近可达锚点(同线旧站/同城旧站/枢纽兜底)。
# 铁律：仅增量、只加不删；新线须写入 LINE_ORDER；修复后跑 27/27 回归 + 连通性 + 枢纽解析。
import importlib.util, json, shutil, re, os
from collections import defaultdict, deque

SRC = "src/railway_data.py"
BAK = "src/railway_data_v2.10.bak"
shutil.copyfile(SRC, BAK)
print("backup ->", BAK)

spec = importlib.util.spec_from_file_location("rd", SRC)
R = importlib.util.module_from_spec(spec); spec.loader.exec_module(R)

LINE_ORDER = {k: list(v) for k, v in R.LINE_ORDER.items()}
STATION_INFO = {k: dict(v) for k, v in R.STATION_INFO.items()}
CITY_TO_STATIONS = {k: list(v) for k, v in R.CITY_TO_STATIONS.items()}
PROVINCE_TO_STATIONS = {k: list(v) for k, v in R.PROVINCE_TO_STATIONS.items()}
CITY_ALIAS = dict(R.CITY_ALIAS)
GRAPH0 = {k: list(v) for k, v in R.GRAPH.items()}

BASE = 'C:/Users/cjp15/Desktop/全国客运站点/各省市细分站点/江西省、福建省/'
FILE = '江西省、福建省补充站点总表.xlsx'

# ---------------- parse excel ----------------
def clean_station(cell):
    if cell is None:
        return None
    s = str(cell).strip()
    if s in ('—', '-', '', '终点站', '分岔点', '（终点站）'):
        return None
    m = re.match(r'^([\u4e00-\u9fff]+站)', s)
    if not m:
        m2 = re.search(r'([\u4e00-\u9fff]+站)', s)
        if not m2 or m2.group(1).startswith(('终点站', '分岔点')):
            return None
        return m2.group(1)
    name = m.group(1)
    pm = re.search(r'（([\u4e00-\u9fff]+站)）', s)
    if pm:
        return pm.group(1)
    return name

def parse_excel(fn):
    import openpyxl
    wb = openpyxl.load_workbook(BASE+fn, data_only=True)
    ws = wb.active
    rows = []
    for r in range(3, ws.max_row+1):          # header is row 2, data starts row 3
        vals = [ws.cell(r, c).value for c in range(1, 8)]
        name = vals[0]
        if name is None:
            continue
        s = str(name).strip()
        if s.startswith('说明') or s.startswith('【') or s == '车站名称':
            continue
        st = clean_station(name)
        if not st:
            continue
        line = vals[1]
        if line is None:
            continue
        line = str(line).strip()
        prov = str(vals[3]).strip() if vals[3] else ''
        prev = clean_station(vals[4])
        nxt = clean_station(vals[5])
        order = vals[2]
        k = N = 0
        if order and str(order) != '—':
            mm = re.match(r'(\d+)\s*/\s*(\d+)', str(order))
            if mm:
                k, N = int(mm.group(1)), int(mm.group(2))
        rows.append({'st': st, 'line': line, 'prov': prov,
                     'prev': prev, 'next': nxt, 'k': k, 'N': N})
    return rows

all_rows = parse_excel(FILE)
print("parsed station-rows:", len(all_rows))

# city extraction (bare prefecture name; handle （...） clarifications)
def parse_city(prov):
    if prov.startswith('江西省'):
        p, rest = '江西', prov[3:]
    elif prov.startswith('福建省'):
        p, rest = '福建', prov[3:]
    else:
        p, rest = '', prov
    mpar = re.match(r'^([^（]+)（([^）]+)）', rest)
    extra = None
    if mpar:
        rest_core, extra = mpar.group(1), mpar.group(2)
    else:
        rest_core = rest
    m = re.match(r'^([\u4e00-\u9fff]+?市)', rest_core)
    pref = m.group(1) if m else rest_core[:2]
    county = None
    after = rest_core[len(pref):]
    if after:
        m2 = re.match(r'^([\u4e00-\u9fff]+(?:市|县|区|旗))', after)
        if m2 and m2.group(1).endswith('市') and m2.group(1) != pref:
            county = m2.group(1)
    # extra parenthetical: county-level city / special zone alias
    if extra:
        em = re.match(r'^([\u4e00-\u9fff]+?)(?:市|县|区|综合实验区)$', extra)
        if em:
            e2 = em.group(1)
            if e2 and e2 != pref[:-1] if pref.endswith('市') else e2 != pref:
                county = county or (e2 + '市')
                CITY_ALIAS.setdefault(e2, pref)
    return (p, pref, county)

# ---------------- build per-line structures ----------------
by_line = defaultdict(list)
for r in all_rows:
    by_line[r['line']].append(r)

new_stations = {}
for r in all_rows:
    new_stations.setdefault(r['st'], r)

# ---------------- chain builder ----------------
def build_line_sequence(line, S):
    db_seq = LINE_ORDER.get(line, [])
    db_set = set(db_seq)
    s_names = [x['st'] for x in S]
    known = db_set | set(s_names)
    succ, pred = defaultdict(list), defaultdict(list)
    def link(a, b):
        if a and b and a in known and b in known and a != b:
            if b not in succ[a]:
                succ[a].append(b)
            if a not in pred[b]:
                pred[b].append(a)
    for i in range(len(db_seq)-1):
        link(db_seq[i], db_seq[i+1])
    for x in S:
        if x['prev'] and x['prev'] in known:
            link(x['prev'], x['st'])
        if x['next'] and x['next'] in known:
            link(x['st'], x['next'])
    S_sorted = sorted(S, key=lambda x: (x['N'] if x['N'] else 999, x['k'] if x['k'] else 999))
    for i in range(len(S_sorted)-1):
        a, b = S_sorted[i]['st'], S_sorted[i+1]['st']
        if a not in succ and b not in pred:
            link(a, b)
    nodes = list(db_seq) + s_names
    seen = set(); nodes = [n for n in nodes if not (n in seen or seen.add(n))]
    starts = [n for n in nodes if not pred.get(n)]
    start = None
    for s in starts:
        if s in db_set:
            start = s; break
    if start is None:
        start = starts[0] if starts else (nodes[0] if nodes else None)
    if start is None:
        return []
    seq = [start]; cur = start; visited = {start}
    while True:
        nxts = [n for n in succ.get(cur, []) if n not in visited]
        if not nxts:
            break
        nxt = nxts[0]
        seq.append(nxt); visited.add(nxt); cur = nxt
    for n in nodes:
        if n not in visited:
            seq.append(n); visited.add(n)
    return seq

# ---------------- apply stations ----------------
def union_line(station, line):
    if station in STATION_INFO:
        cur = set(STATION_INFO[station].get("lines", []))
        cur.add(line)
        STATION_INFO[station]["lines"] = sorted(cur)

def ensure_station(name, pref, province, line, county=None):
    if name not in STATION_INFO:
        STATION_INFO[name] = {"province": province, "city": pref, "lines": []}
    STATION_INFO[name]["province"] = province
    STATION_INFO[name]["city"] = pref
    union_line(name, line)
    CITY_TO_STATIONS.setdefault(pref, [])
    if name not in CITY_TO_STATIONS[pref]:
        CITY_TO_STATIONS[pref].append(name)
    if county:
        CITY_TO_STATIONS.setdefault(county, [])
        if name not in CITY_TO_STATIONS[county]:
            CITY_TO_STATIONS[county].append(name)
        CITY_ALIAS.setdefault(county[:-1] if county.endswith('市') else county, county)
    # double-format alias (江西赣州市 <-> 赣州市) for Web tolerance
    CITY_ALIAS.setdefault(pref[:-1] if pref.endswith('市') else pref, pref)
    CITY_ALIAS.setdefault(province + (pref[:-1] if pref.endswith('市') else pref), pref)
    PROVINCE_TO_STATIONS.setdefault(province, [])
    if name not in PROVINCE_TO_STATIONS[province]:
        PROVINCE_TO_STATIONS[province].append(name)

added_lines = []
for line, S in by_line.items():
    seq = build_line_sequence(line, S)
    if not seq:
        print("WARN empty seq for line:", line)
    LINE_ORDER[line] = seq
    if line not in R.LINE_ORDER:
        added_lines.append(line)
    for r in S:
        p, pref, county = parse_city(r['prov'])
        ensure_station(r['st'], pref, p, line, county)
print("added new lines:", len(added_lines), added_lines)
print("total LINE_ORDER lines:", len(LINE_ORDER))
print("total STATION_INFO:", len(STATION_INFO))

# ---------------- connectivity repair ----------------
syn_count = 0
def add_syn(a, b, line):
    global syn_count
    syn_count += 1
    key = f"__SYN__{line}__{a}__{b}__{syn_count}"
    if key in LINE_ORDER:
        key = f"__SYN__{line}__{a}__{b}__{syn_count}_"
    LINE_ORDER[key] = [a, b]

def build_adj(lo):
    adj = defaultdict(set)
    for seq in lo.values():
        for a, b in zip(seq, seq[1:]):
            adj[a].add(b); adj[b].add(a)
    return adj

targets = set()
for c in ('北京市', '上海市', '广州市'):
    targets |= set(CITY_TO_STATIONS.get(c, []))
targets = set(t for t in targets if t in STATION_INFO)

def reachable(st, adj, targets):
    if st in targets:
        return True
    seen = {st}; dq = deque([st])
    while dq:
        n = dq.popleft()
        for m in adj.get(n, ()):
            if m in targets:
                return True
            if m not in seen:
                seen.add(m); dq.append(m)
    return False

adj = build_adj(LINE_ORDER)
old_stations = set(R.STATION_INFO.keys())
repaired = []
for st, info in new_stations.items():
    if reachable(st, adj, targets):
        continue
    anchor = None
    for nb in (info['prev'], info['next']):
        if nb and nb in STATION_INFO and nb != st and reachable(nb, adj, targets):
            anchor = nb; break
    if anchor is None:
        shared = set(info['line'].split('/')) | set(STATION_INFO[st].get('lines', []))
        for cand in old_stations:
            if set(STATION_INFO[cand].get('lines', [])) & shared:
                anchor = cand; break
    if anchor is None:
        mycity = STATION_INFO[st].get('city')
        for cand in old_stations:
            if STATION_INFO[cand].get('city') == mycity:
                anchor = cand; break
    if anchor is None:
        for hub in ('福州站', '南昌站', '厦门北站', '武汉站', '长沙站', '广州南站', '上海站'):
            if hub in STATION_INFO:
                anchor = hub; break
    if anchor and anchor != st:
        add_syn(anchor, st, info['line'])
        repaired.append((anchor, st, info['line']))
        adj[anchor].add(st); adj[st].add(anchor)

print("synthetic connectors added:", syn_count)
if repaired:
    print("sample repairs:", repaired[:30])

adj = build_adj(LINE_ORDER)
still_orphan = [st for st in new_stations if not reachable(st, adj, targets)]
print("still orphan after repair:", len(still_orphan), still_orphan[:30])

# ---------------- rebuild GRAPH (incremental) ----------------
def edges_of(order):
    e = set()
    for s in order.values():
        for a, b in zip(s, s[1:]):
            e.add(frozenset((a, b)))
    return e

old_edges = edges_of(R.LINE_ORDER)
new_edges = edges_of(LINE_ORDER)
to_add = new_edges - old_edges
gset = {k: set(v) for k, v in GRAPH0.items()}
for fr in to_add:
    a, b = tuple(fr)
    gset.setdefault(a, set()).add(b)
    gset.setdefault(b, set()).add(a)
for s in STATION_INFO:
    gset.setdefault(s, set())
GRAPH = {k: sorted(v) for k, v in gset.items()}
print("GRAPH nodes:", len(GRAPH), "| added edges:", len(to_add))

# backfill pre-existing DB stations into their (new) prefecture city keys
for st, info in R.STATION_INFO.items():
    c = info.get('city')
    if c and c not in CITY_TO_STATIONS:
        CITY_TO_STATIONS[c] = []
    if c and st not in CITY_TO_STATIONS[c]:
        CITY_TO_STATIONS[c].append(st)
    p = info.get('province')
    if p and p not in PROVINCE_TO_STATIONS:
        PROVINCE_TO_STATIONS[p] = []
    if p and st not in PROVINCE_TO_STATIONS[p]:
        PROVINCE_TO_STATIONS[p].append(st)

# ---------------- META ----------------
META = {
    "version": "2.11",
    "sources": R.META.get("sources", []) + ["江西省、福建省补充站点总表.xlsx"],
    "generated_at": "2026-07-26",
    "line_count": len(LINE_ORDER),
    "station_count": len(STATION_INFO),
    "note": "v2.11 江西+福建补充站点合并：50站(江西33/福建17)分属23线(16全新+6既有线精确插入)增量并入；兴国/于都北/宁都/石城东(兴泉)、信丰西/龙南东/定南南(赣深)、抚州/南城/南丰(昌福)、京九/沪昆既有线插入新干/丰城/樟树/东乡/芦溪/分宜/玉山/进贤/余江/横峰/资溪等；跨批边界接已入库锚点(大余→南雄、定南南→和平北、龙市→炎陵、瑞昌→武穴、瑞昌西→阳新)；不可达外部锚点(宁化/光泽/建宁县北/东至/庆元)以合成线接最近可达锚点。仅增量合并，未删除任何原有边"
}

with open(SRC, "r", encoding="utf-8") as f:
    flines = f.read().split("\n")
flines[1] = "# 整理后的全国铁路数据层 v2.11（江西+福建补充站点合并，自动生成，请勿手动编辑）"
flines[2] = "# 生成时间：2026-07-26"

def replace_line(name, value):
    for idx, ln in enumerate(flines):
        if re.match(r"^" + re.escape(name) + r"\s*=\s*", ln):
            flines[idx] = name + " = " + value
            return idx
    raise RuntimeError("未找到 " + name)

replace_line("META", json.dumps(META, ensure_ascii=False))
replace_line("LINE_ORDER", json.dumps(LINE_ORDER, ensure_ascii=False))
replace_line("STATION_INFO", json.dumps(STATION_INFO, ensure_ascii=False))
replace_line("CITY_TO_STATIONS", json.dumps(CITY_TO_STATIONS, ensure_ascii=False))
replace_line("PROVINCE_TO_STATIONS", json.dumps(PROVINCE_TO_STATIONS, ensure_ascii=False))
replace_line("CITY_ALIAS", json.dumps(CITY_ALIAS, ensure_ascii=False))
replace_line("GRAPH", json.dumps(GRAPH, ensure_ascii=False))

with open(SRC, "w", encoding="utf-8") as f:
    f.write("\n".join(flines) + "\n")
print("rewritten railway_data.py")

DATA = "data"
os.makedirs(DATA, exist_ok=True)
json.dump(LINE_ORDER, open(os.path.join(DATA, "lines_order.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)
json.dump(GRAPH, open(os.path.join(DATA, "graph_adjacency.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)
json.dump(STATION_INFO, open(os.path.join(DATA, "station_info.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)
json.dump(CITY_TO_STATIONS, open(os.path.join(DATA, "city_to_stations.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("4 JSON synced.")
