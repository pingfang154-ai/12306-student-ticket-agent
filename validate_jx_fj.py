# -*- coding: utf-8 -*-
import importlib.util, json
from collections import deque

def load(p):
    import importlib.machinery
    nm = "rdmod_x"
    loader = importlib.machinery.SourceFileLoader(nm, p)
    spec = importlib.util.spec_from_loader(nm, loader)
    m = importlib.util.module_from_spec(spec)
    loader.exec_module(m)
    return m

cur = load("src/railway_data.py")
bak = load("src/railway_data_v2.10.bak")

print("=== META ===")
print("cur version:", cur.META.get("version"), "| lines:", len(cur.LINE_ORDER), "stations:", len(cur.STATION_INFO), "graph nodes:", len(cur.GRAPH))
print("bak version:", bak.META.get("version"), "| lines:", len(bak.LINE_ORDER), "stations:", len(bak.STATION_INFO))

# new stations this batch (in cur not in bak)
new = [s for s in cur.STATION_INFO if s not in bak.STATION_INFO]
print("\n=== NEW stations count:", len(new))

# 1) resolve non-empty for all 50 expected
import openpyxl
BASE='C:/Users/cjp15/Desktop/全国客运站点/各省市细分站点/江西省、福建省/'
wb = openpyxl.load_workbook(BASE+'江西省、福建省补充站点总表.xlsx', data_only=True)
ws = wb.active
expected=[]
for r in range(3, ws.max_row+1):
    name = ws.cell(r,1).value
    if name is None: continue
    s=str(name).strip()
    if s.startswith('说明') or s.startswith('【') or s=='车站名称': continue
    import re
    def cs(c):
        if c is None: return None
        t=str(c).strip()
        if t in ('—','-','',): return None
        mm=re.match(r'^([\u4e00-\u9fff]+站)', t)
        return mm.group(1) if mm else None
    st=cs(name)
    if st: expected.append(st)
print("expected stations from excel:", len(expected))
empty=[s for s in expected if not cur.resolve_location(s)]
print("resolve EMPTY for expected:", len(empty), empty[:20])

# 2) new cities reachable to 北上广
def adj_of(rd):
    from collections import defaultdict
    a=defaultdict(set)
    for seq in rd.LINE_ORDER.values():
        for x,y in zip(seq, seq[1:]):
            a[x].add(y); a[y].add(x)
    return a
adj=adj_of(cur)
targets=set()
for c in ('北京市','上海市','广州市'):
    targets|=set(cur.CITY_TO_STATIONS.get(c,[]))
targets={t for t in targets if t in cur.STATION_INFO}
def reach(st):
    if st in targets: return True
    seen={st}; dq=deque([st])
    while dq:
        n=dq.popleft()
        for m in adj.get(n,()):
            if m in targets: return True
            if m not in seen: seen.add(m); dq.append(m)
    return False
new_cities=set()
for s in new:
    new_cities.add(cur.STATION_INFO[s].get('city'))
unreached_cities=[]
for c in sorted(new_cities):
    sts=cur.CITY_TO_STATIONS.get(c,[])
    if sts and not any(reach(x) for x in sts):
        unreached_cities.append(c)
print("\n=== new prefecture cities:", len(new_cities), sorted(new_cities))
print("unreached cities (to 北上广):", len(unreached_cities), unreached_cities)

# per new station reachability
unreached_st=[s for s in new if not reach(s)]
print("unreached NEW stations:", len(unreached_st), unreached_st[:30])

# 3) hub parsing compare
def hubs():
    try:
        return json.load(open("data/hub_stations.json",encoding="utf-8"))
    except: return []
hubs_list=hubs()
def resolve_count(rd):
    ok=0; bad=[]
    for h in hubs_list:
        nm=h.get('name') if isinstance(h,dict) else h
        if rd.resolve_location(nm): ok+=1
        else: bad.append(nm)
    return ok,len(hubs_list),bad
ok,total,bad=resolve_count(cur)
print("\n=== HUB parsing (cur v2.11):", ok,"/",total, "unparsed:", bad)
okb,tb,badb=resolve_count(bak)
print("HUB parsing (v2.10 bak):", okb,"/",tb,"unparsed:", badb)
print("newly unparsed vs bak:", [x for x in bad if x in badb])
