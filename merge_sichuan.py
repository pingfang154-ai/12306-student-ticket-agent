# -*- coding: utf-8 -*-
"""增量合并：四川省重庆市补充站点总表 -> 第三版 railway_data.py (v2.4 -> v2.5)"""
import importlib.util, sys, os, json, re

BASE = r"C:\Users\cjp15\Desktop\全国客运站点\交接文件夹第三版（含一、二合并）"
SRC = os.path.join(BASE, "src")
XLSX = r"C:\Users\cjp15\Desktop\全国客运站点\各省市细分站点\四川省、重庆市\四川省重庆市补充站点总表.xlsx"
sys.path.insert(0, SRC)
spec = importlib.util.spec_from_file_location("rd", os.path.join(SRC, "railway_data.py"))
R = importlib.util.module_from_spec(spec); spec.loader.exec_module(R)

import openpyxl
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active
# header at row 2: 车站名称/所属线路/所属省市/前一站/后一站/备注
rows = []
for r in range(3, ws.max_row + 1):
    name = ws.cell(row=r, column=1).value
    if not name or str(name).strip() == "":
        continue
    if str(name).strip().startswith("说明"):
        break
    line = ws.cell(row=r, column=2).value
    prov_city = ws.cell(row=r, column=3).value
    prev = ws.cell(row=r, column=4).value
    nxt = ws.cell(row=r, column=5).value
    note = ws.cell(row=r, column=6).value
    rows.append((str(name).strip(), str(line).strip() if line else "",
                 str(prov_city).strip() if prov_city else "",
                 str(prev).strip() if prev else "",
                 str(nxt).strip() if nxt else "",
                 str(note).strip() if note else ""))

def norm(s):
    if not s:
        return ""
    s = s.strip()
    s = re.sub(r"[（(].*?[)）]", "", s)  # drop parentheticals e.g. (金江站)/(湖北省)
    s = s.replace("站", "").strip()
    return s + "站"

def extract_city(prov_city):
    """Return (province_short, city_full)."""
    pc = prov_city.strip()
    if pc.startswith("四川省"):
        prov = "四川"
        rest = pc[len("四川省"):]
    elif pc.startswith("重庆市"):
        prov = "重庆"
        rest = pc[len("重庆市"):]
    else:
        # fallback
        prov = "四川"
        rest = pc
    # take first administrative token ending 市/州/区/县/自治县
    m = re.search(r"([\u4e00-\u9fa5]{1,10}?(?:市|州|区|县|自治县|土家族自治县|苗族土家族自治县))", rest)
    if m:
        city = m.group(1)
    else:
        city = rest.split("市")[0] + "市" if "市" in rest else rest
    return prov, city

# Build SUPP records
SUPP = {}  # station -> dict
for name, line, pc, prev, nxt, note in rows:
    st = norm(name)
    if st == "开江南站":
        continue  # 在建未开通，跳过（避免误判可购票）
    prov, city = extract_city(pc)
    SUPP[st] = {
        "raw_line": line,
        "prov": prov,
        "city": city,
        "prev": norm(prev) if prev and "起点" not in prev and "终点" not in prev else "",
        "next": norm(nxt) if nxt and "起点" not in nxt and "终点" not in nxt else "",
        "note": note,
        "passenger": ("停办" not in note) and ("不办客运" not in note) and ("暂停客运" not in note),
    }

# LINE_MAP: supplement raw line -> DB LINE_ORDER key (create new if needed)
LINE_MAP = {
    "成昆铁路（攀枝花沿江段，现峨攀支线）": "成昆铁路（含复线）",
    "成昆铁路（攀枝花沿江段）": "成昆铁路（含复线）",
    "成昆铁路（峨眉—普雄段）": "成昆铁路（含复线）",
    "成昆铁路（普雄—西昌段）": "成昆铁路（含复线）",
    "成昆铁路（西昌—攀枝花段）": "成昆铁路（含复线）",
    "新成昆铁路（成昆铁路复线）": "新成昆铁路",
    "达成铁路": "达成铁路",
    "襄渝铁路": "襄渝铁路",
    "兰渝铁路": "兰渝铁路",
    "兰渝铁路高南支线（南高支线）": "兰渝铁路",
    "渝怀铁路（普速线）": "渝怀铁路",
    "宝成铁路": "宝成铁路",
    "川青铁路（成兰铁路镇江关至黄胜关段）": "川青铁路",
    "川青铁路（成兰铁路）": "川青铁路",
    "西成高速铁路": "西成高速铁路",
    "渝昆高速铁路（渝宜段）": "渝昆高速铁路（渝宜段）",
    "渝昆高铁（渝宜段）": "渝昆高速铁路（渝宜段）",
    "渝昆高铁（川渝段）": "渝昆高速铁路（渝宜段）",
    "渝昆高铁（川渝段）/绵泸高铁（内自泸段）": "渝昆高速铁路（渝宜段）",
    "成自宜高铁（川南城际铁路自宜段）": "成自宜高铁",
    "郑渝高速铁路": "郑渝高速铁路",
    "郑渝高铁": "郑渝高速铁路",
    "渝利铁路": "渝利铁路",
    "内昆铁路（内六铁路）/宜珙铁路": "内昆铁路（内六段）",
    "成雅铁路（川藏铁路成雅段）": "成雅铁路",
    "川藏铁路成雅段（成雅铁路）": "成雅铁路",
    "隆黄铁路（叙毕铁路段）": "隆黄铁路（叙毕段）",
    "绵泸高铁（内自泸段）": "绵泸高铁（内自泸段）",
    "巴南高铁（巴中东至南充段）": "巴南高铁",
    "巴南高铁": "巴南高铁",
    "巴南高铁（巴中东至南充段） / 巴达铁路": "巴达铁路",
    "巴达铁路": "巴达铁路",
    "广巴铁路 / 巴达铁路": "巴达铁路",
    "渝厦高铁（重庆东至黔江段）": "渝厦高铁（重庆段）",
    "南涪铁路（涪三线）": "南涪铁路",
    # additional raw variants seen in data
    "成昆铁路（攀枝花沿江段，现峨攀支线终点）": "成昆铁路（含复线）",
    "兰渝铁路（广渝段主线）": "兰渝铁路",
    "兰渝铁路高南支线（接轨襄渝铁路）": "兰渝铁路",
    "川青铁路（成兰铁路镇江关至黄胜关段）": "川青铁路",
}
# any unmapped raw line -> create a sanitized key
for st, rec in SUPP.items():
    rl = rec["raw_line"]
    if rl not in LINE_MAP:
        key = re.sub(r"[（(].*?[)）]", "", rl).strip()
        LINE_MAP[rl] = key
        print("WARN new line key:", key, "<-", rl)

# group stations by DB line key
by_line = {}
for st, rec in SUPP.items():
    key = LINE_MAP[rec["raw_line"]]
    by_line.setdefault(key, []).append(st)

# Build directed edges per line: db edges + supp edges + connectors
# connectors = referenced prev/next not in SUPP and not in DB STATION_INFO
db_stations = set(R.STATION_INFO.keys())

# directed adjacency: line_key -> {node: set(successors)}
adj = {}
rev_adj = {}
modified_lines = set(by_line.keys())

def add_edge(key, a, b):
    adj.setdefault(key, {}).setdefault(a, set()).add(b)
    rev_adj.setdefault(key, {}).setdefault(b, set()).add(a)

for key in modified_lines:
    seq = R.LINE_ORDER.get(key, [])
    for a, b in zip(seq, seq[1:]):
        add_edge(key, a, b)

# supp edges
for key, sts in by_line.items():
    for st in sts:
        rec = SUPP[st]
        p, n = rec["prev"], rec["next"]
        if p:
            add_edge(key, p, st)
        if n:
            add_edge(key, st, n)

# collect connector nodes (referenced but not supp & not db)
connectors = {}
for key, sts in by_line.items():
    for st in sts:
        rec = SUPP[st]
        for ref in (rec["prev"], rec["next"]):
            if ref and ref not in SUPP and ref not in db_stations:
                connectors[ref] = key

# assign connector cities: use the referencing supp station's city (best-effort)
connector_city = {}
for key, sts in by_line.items():
    for st in sts:
        rec = SUPP[st]
        for ref in (rec["prev"], rec["next"]):
            if ref in connectors:
                connector_city.setdefault(ref, (rec["prov"], rec["city"]))
# connectors that are themselves referenced multiple times keep first

# MANUAL extra directed edges to bridge chains to DB anchors
MANUAL_EDGES = [
    # 成昆 普速
    ("成昆铁路（含复线）", "西昌站", "西昌南站"),
    ("成昆铁路（含复线）", "德昌站", "攀枝花站"),
    ("成昆铁路（含复线）", "花棚子站", "永仁站"),
    ("成昆铁路（含复线）", "西昌站", "冕山站"),
    ("成昆铁路（含复线）", "尼日站", "峨眉站"),
    # 襄渝
    ("襄渝铁路", "重庆西站", "北碚站"),
    ("襄渝铁路", "北碚站", "广安站"),
    ("襄渝铁路", "广安站", "华蓥站"),
    ("襄渝铁路", "宣汉站", "毛坝站"),
    ("襄渝铁路", "毛坝站", "万源站"),
    # 兰渝高南支线 (广安南/岳池/高兴)
    ("兰渝铁路", "南充东站", "岳池站"),
    ("兰渝铁路", "岳池站", "广安南站"),
    ("兰渝铁路", "广安南站", "高兴站"),
    ("兰渝铁路", "武胜站", "合川站"),  # already? ensure
    # 达成
    ("达成铁路", "南充站", "蓬溪站"),
    ("达成铁路", "蓬溪站", "遂宁站"),
    ("达成铁路", "蓬溪站", "大英东站"),
    ("达成铁路", "大英东站", "成都东站"),
    # 宝成 (new line): connect via 广元/绵阳
    ("宝成铁路", "广元站", "朝天南站"),
    ("宝成铁路", "朝天南站", "观音坝站"),
    ("宝成铁路", "绵阳站", "黄许镇站"),
    ("宝成铁路", "黄许镇站", "罗江站"),
    # 川青
    ("川青铁路", "镇江关站", "松潘站"),
    ("川青铁路", "松潘站", "黄龙九寨站"),
    ("川青铁路", "黄龙九寨站", "黄胜关站"),
    # 西成
    ("西成高速铁路", "广元站", "朝天站"),
    ("西成高速铁路", "朝天站", "宁强南站"),
    # 渝昆
    ("渝昆高速铁路（渝宜段）", "重庆西站", "泸州东站"),
    ("渝昆高速铁路（渝宜段）", "泸州东站", "泸州站"),
    ("渝昆高速铁路（渝宜段）", "泸州站", "南溪站"),
    ("渝昆高速铁路（渝宜段）", "南溪站", "宜宾东站"),
    # 成自宜
    ("成自宜高铁", "自贡站", "沿滩站"),
    ("成自宜高铁", "沿滩站", "南溪北站"),
    ("成自宜高铁", "南溪北站", "宜宾东站"),
    # 内昆
    ("内昆铁路（内六段）", "宜宾站", "宜宾北站"),
    ("内昆铁路（内六段）", "宜宾北站", "翠屏站"),
    ("内昆铁路（内六段）", "翠屏站", "宜宾南站"),
    # 成雅 (new): connect to 成都 network via 朝阳湖->成蒲? bridge to 成都西/成都
    ("成雅铁路", "朝阳湖站", "名山站"),
    ("成雅铁路", "名山站", "雅安站"),
    # 隆黄叙毕 (new): connect 叙永北->泸州/毕节. Use 泸州 as hub.
    ("隆黄铁路（叙毕段）", "兴文南站", "金桂站"),
    ("隆黄铁路（叙毕段）", "金桂站", "白腊站"),
    ("隆黄铁路（叙毕段）", "白腊站", "叙永北站"),
    ("隆黄铁路（叙毕段）", "叙永北站", "龙凤镇站"),
    # 绵泸内自泸 (new): connect via 自贡/泸州
    ("绵泸高铁（内自泸段）", "内江东站", "白马北站"),
    ("绵泸高铁（内自泸段）", "白马北站", "自贡站"),
    ("绵泸高铁（内自泸段）", "自贡站", "富顺站"),
    ("绵泸高铁（内自泸段）", "富顺站", "泸县站"),
    ("绵泸高铁（内自泸段）", "泸县站", "泸州站"),
    # 巴南高铁 (new): connect via 南充北
    ("巴南高铁", "南充北站", "蓬安西站"),
    ("巴南高铁", "蓬安西站", "仪陇站"),
    ("巴南高铁", "仪陇站", "马鞍站"),
    ("巴南高铁", "马鞍站", "巴中西站"),
    ("巴南高铁", "巴中西站", "巴中东站"),
    # 巴达铁路 (new): connect 巴中->达州
    ("巴达铁路", "达州站", "石梯站"),
    ("巴达铁路", "石梯站", "何家坪站"),
    ("巴达铁路", "何家坪站", "平昌站"),
    ("巴达铁路", "平昌站", "曾口站"),
    ("巴达铁路", "曾口站", "巴中东站"),
    ("巴达铁路", "巴中东站", "巴中站"),
    # 渝怀
    ("渝怀铁路", "武隆站", "彭水站"),
    ("渝怀铁路", "彭水站", "黔江站"),
    ("渝怀铁路", "武隆站", "白马站"),
    # 渝利
    ("渝利铁路", "重庆北站", "涪陵北站"),
    ("渝利铁路", "涪陵北站", "丰都站"),
    ("渝利铁路", "丰都站", "石柱县站"),
    # 郑渝
    ("郑渝高速铁路", "重庆北站", "云阳站"),
    ("郑渝高速铁路", "云阳站", "奉节站"),
    ("郑渝高速铁路", "奉节站", "巫山站"),
    ("郑渝高速铁路", "巫山站", "巴东站"),
    # 渝厦重庆段 (new)
    ("渝厦高铁（重庆段）", "南川北站", "水江西站"),
    ("渝厦高铁（重庆段）", "水江西站", "武隆南站"),
    ("渝厦高铁（重庆段）", "武隆南站", "彭水西站"),
    ("渝厦高铁（重庆段）", "彭水西站", "黔江站"),
    # 南涪 (new): connect 水江 to 南川
    ("南涪铁路", "南川站", "水江站"),
    # 兰渝 武胜 already
]
for key, a, b in MANUAL_EDGES:
    if key not in modified_lines and key not in R.LINE_ORDER:
        modified_lines.add(key)
    add_edge(key, a, b)

# Topological sort per line (Kahn on directed graph)
def topo_sort(key):
    a = adj.get(key, {})
    indeg = {n: 0 for n in set(list(a.keys()) + [x for s in a.values() for x in s])}
    for u in a:
        for v in a[u]:
            indeg[v] = indeg.get(v, 0) + 1
    # roots = indeg 0
    from collections import deque
    q = deque([n for n in indeg if indeg[n] == 0])
    order = []
    indeg_copy = dict(indeg)
    while q:
        u = q.popleft()
        order.append(u)
        for v in a.get(u, ()):
            indeg_copy[v] -= 1
            if indeg_copy[v] == 0:
                q.append(v)
    # any leftover (cycle) append
    if len(order) < len(indeg):
        for n in indeg:
            if n not in order:
                order.append(n)
    return order

NEW_LINE_ORDER = dict(R.LINE_ORDER)
for key in modified_lines:
    seq = topo_sort(key)
    # dedupe preserving order
    seen = set(); out = []
    for s in seq:
        if s not in seen:
            seen.add(s); out.append(s)
    NEW_LINE_ORDER[key] = out

# ---- Patch: guarantee every intended adjacency edge exists as a consecutive pair ----
# Intended edges = original db consecutive pairs (for modified lines) + supp prev/next + manual edges
INTENDED = {k: set() for k in modified_lines}
for key in modified_lines:
    seq = R.LINE_ORDER.get(key, [])
    for a, b in zip(seq, seq[1:]):
        INTENDED[key].add((a, b))
for key, sts in by_line.items():
    for st in sts:
        rec = SUPP[st]
        if rec["prev"]:
            INTENDED[key].add((rec["prev"], st))
        if rec["next"]:
            INTENDED[key].add((st, rec["next"]))
for key, a, b in MANUAL_EDGES:
    if key in INTENDED:
        INTENDED[key].add((a, b))

def is_consecutive(seq, a, b):
    for i in range(len(seq) - 1):
        if seq[i] == a and seq[i + 1] == b:
            return True
    return False

# (synthetic bridge application moved below, after add_station is defined)

# ---- Update STATION_INFO / CITY / PROVINCE ----
STATION_INFO = {k: dict(v) for k, v in R.STATION_INFO.items()}
CITY_TO_STATIONS = {k: list(v) for k, v in R.CITY_TO_STATIONS.items()}
PROVINCE_TO_STATIONS = {k: list(v) for k, v in R.PROVINCE_TO_STATIONS.items()}
CITY_ALIAS = dict(R.CITY_ALIAS)

def add_station(st, prov, city, line_key):
    if st in STATION_INFO:
        info = STATION_INFO[st]
        if line_key not in info.get("lines", []):
            info["lines"] = info.get("lines", []) + [line_key]
    else:
        STATION_INFO[st] = {"province": prov, "city": city, "lines": [line_key]}
    # city map
    CITY_TO_STATIONS.setdefault(city, [])
    if st not in CITY_TO_STATIONS[city]:
        CITY_TO_STATIONS[city].append(st)
    PROVINCE_TO_STATIONS.setdefault(prov, [])
    if st not in PROVINCE_TO_STATIONS[prov]:
        PROVINCE_TO_STATIONS[prov].append(st)
    # alias short->full (strip trailing 市/州/区/县/自治县)
    short = re.sub(r"(市|州|区|县|自治县|土家族自治县|苗族土家族自治县)$", "", city)
    if short and short not in CITY_ALIAS and short != city:
        CITY_ALIAS[short] = city

# add supplement stations
for st, rec in SUPP.items():
    key = LINE_MAP[rec["raw_line"]]
    add_station(st, rec["prov"], rec["city"], key)
# add connectors
for c, key in connectors.items():
    prov, city = connector_city.get(c, ("四川", c))
    add_station(c, prov, city, key)

# ---- Apply synthetic bridges: guarantee every intended adjacency is a consecutive pair ----
syn_count = 0
for key in modified_lines:
    seq = NEW_LINE_ORDER[key]
    for a, b in INTENDED[key]:
        if a in seq and b in seq and not is_consecutive(seq, a, b):
            syn_key = f"__SYN__{key}__{a}__{b}"
            NEW_LINE_ORDER[syn_key] = [a, b]
            syn_count += 1
            for ep in (a, b):
                if ep not in STATION_INFO and ep not in SUPP:
                    prov, city = connector_city.get(ep, ("四川", ep))
                    add_station(ep, prov, city, key)
print("synthetic bridge lines added:", syn_count)

# ---- Rebuild GRAPH incrementally for modified lines ----
GRAPH = {k: set(v) for k, v in R.GRAPH.items()}
def line_edges(seq):
    e = set()
    for a, b in zip(seq, seq[1:]):
        e.add((a, b)); e.add((b, a))
    return e
old_edges = set()
for key in modified_lines:
    if key in R.LINE_ORDER:
        old_edges |= line_edges(R.LINE_ORDER[key])
new_edges = set()
for key in modified_lines:
    new_edges |= line_edges(NEW_LINE_ORDER[key])
# remove old, add new
for a, b in old_edges:
    if a in GRAPH and b in GRAPH.get(a, ()):
        GRAPH[a].discard(b)
        GRAPH[b].discard(a)
for a, b in new_edges:
    GRAPH.setdefault(a, set()).add(b)
    GRAPH.setdefault(b, set()).add(a)

# safety: every GRAPH node must exist in STATION_INFO (else resolver can't see it)
for node in list(GRAPH.keys()):
    if node not in STATION_INFO:
        prov, city = connector_city.get(node, ("四川", node))
        add_station(node, prov, city, "__orphan__")

# ---- Regenerate railway_data.py ----
with open(os.path.join(SRC, "railway_data.py"), encoding="utf-8") as f:
    orig_all = f.read().splitlines()
head_end = 6  # comments (lines 1-6)
# find helper start marker
hi = next(i for i, l in enumerate(orig_all) if l.strip().startswith("LINE_NAME_ALIAS") or "def resolve_location" in l)
head_lines = orig_all[:head_end]
tail_lines = orig_all[hi:]

import json as _json
def dump(d):
    return _json.dumps(d, ensure_ascii=False)

out = []
out.extend(head_lines)
out.append("META = " + dump(R.META))
out.append("")
out.append("LINE_ORDER = " + dump(NEW_LINE_ORDER))
out.append("")
out.append("STATION_INFO = " + dump(STATION_INFO))
out.append("")
out.append("CITY_TO_STATIONS = " + dump(CITY_TO_STATIONS))
out.append("")
out.append("PROVINCE_TO_STATIONS = " + dump(PROVINCE_TO_STATIONS))
out.append("")
out.append("CITY_ALIAS = " + dump(CITY_ALIAS))
out.append("")
out.append("GRAPH = " + dump({k: sorted(v) for k, v in GRAPH.items()}))
out.append("")
out.extend(tail_lines)
if not out[-1].endswith("\n"):
    out[-1] += "\n"
with open(os.path.join(SRC, "railway_data.py"), "w", encoding="utf-8") as f:
    f.write("\n".join(out))

# ---- Sync JSON ----
def sync(path, d):
    with open(path, "w", encoding="utf-8") as f:
        _json.dump(d, f, ensure_ascii=False)
sync(os.path.join(BASE, "data", "lines_order.json"), {"lines_order": NEW_LINE_ORDER})
sync(os.path.join(BASE, "data", "graph_adjacency.json"), {"graph_adjacency": {k: sorted(v) for k, v in GRAPH.items()}})
sync(os.path.join(BASE, "data", "station_info.json"), {"station_info": STATION_INFO})
sync(os.path.join(BASE, "data", "city_to_stations.json"), {"city_to_stations": CITY_TO_STATIONS})

# ---- Report ----
print("=== MERGE SUMMARY ===")
print("supplement stations added:", len([s for s in SUPP]))
print("connectors added:", len(connectors), list(connectors.keys()))
print("modified lines:", len(modified_lines))
for k in sorted(modified_lines):
    seq = NEW_LINE_ORDER[k]
    print(f"  {k} ({len(seq)}): {' - '.join(seq[:6])}{' ...' if len(seq)>6 else ''}")
print("LINE_ORDER total:", len(NEW_LINE_ORDER))
print("STATION_INFO total:", len(STATION_INFO))
print("GRAPH nodes:", len(GRAPH))
